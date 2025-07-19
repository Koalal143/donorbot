from __future__ import annotations

import asyncio
from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from aiogram import Bot

    from src.repositories.donation import DonationRepository


class ExcelGenerationService:
    def __init__(self, bot: Bot) -> None:
        self.bot = bot

    async def generate_statistics_excel(
        self, organizer_id: int, donation_repository: DonationRepository
    ) -> dict[str, Any]:
        """Генерировать Excel файл со статистикой донорских дней организатора"""
        try:
            # Получаем статистику из БД
            statistics = await donation_repository.get_all_organizer_donor_days(organizer_id)

            if not statistics:
                return {
                    "success": False,
                    "error": "У организатора нет донорских дней в базе данных",
                    "file_content": None,
                }

            # Генерируем Excel файл в отдельном потоке
            file_content = await asyncio.to_thread(self._generate_excel_content, statistics)

            return {
                "success": True,
                "file_content": file_content,
                "records_count": len(statistics),
                "filename": f"statistics_organizer_{organizer_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            }

        except Exception as e:
            return {"success": False, "error": f"Ошибка при генерации файла: {e!s}", "file_content": None}

    def _generate_excel_content(self, statistics: list[dict[str, Any]]) -> BytesIO:
        """Генерировать содержимое Excel файла со статистикой"""
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Статистика донорских дней"

        # Заголовки
        headers = ["ДАТА ДД", "ЦЕНТР КРОВИ", "КОЛИЧЕСТВО ДОНОРОВ", "КОЛИЧЕСТВО РЕГИСТРАЦИЙ"]

        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Записываем данные
        for row_num, stat in enumerate(statistics, 2):
            # Дата донорского дня
            date_str = stat["event_date"].strftime("%d.%m.%Y") if stat["event_date"] else ""
            sheet.cell(row=row_num, column=1, value=date_str)

            # Центр крови (имя организатора)
            sheet.cell(row=row_num, column=2, value=stat["organizer_name"])

            # Количество доноров (подтвержденные донации)
            sheet.cell(row=row_num, column=3, value=stat["confirmed_donations"])

            # Количество регистраций
            sheet.cell(row=row_num, column=4, value=stat["total_registrations"])

        # Автоматическая ширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True
            # Устанавливаем минимальную ширину
            sheet.column_dimensions[column_letter].width = max(15, sheet.column_dimensions[column_letter].width or 0)

        # Сохраняем в BytesIO
        file_content = BytesIO()
        workbook.save(file_content)
        file_content.seek(0)

        return file_content

    async def generate_donors_excel(self, organizer_id: int, donation_repository: DonationRepository) -> dict[str, Any]:
        """Генерировать Excel файл со статистикой доноров (подтвержденные донации)"""
        try:
            # Получаем данные подтвержденных донаций из БД
            donations = await donation_repository.get_confirmed_donations_by_organizer(organizer_id)

            if not donations:
                return {
                    "success": False,
                    "error": "У организатора нет подтвержденных донаций в базе данных",
                    "file_content": None,
                }

            # Генерируем Excel файл в отдельном потоке
            file_content = await asyncio.to_thread(self._generate_donors_excel_content, donations)

            return {
                "success": True,
                "file_content": file_content,
                "records_count": len(donations),
                "filename": f"donors_organizer_{organizer_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            }

        except Exception as e:
            return {"success": False, "error": f"Ошибка при генерации файла: {e!s}", "file_content": None}

    def _generate_donors_excel_content(self, donations: list[dict[str, Any]]) -> BytesIO:
        """Генерировать содержимое Excel файла со статистикой доноров"""
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Статистика доноров"

        # Заголовки
        headers = ["ФИО", "ДАТА", "ОРГАНИЗАТОР (ЦК)"]

        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Записываем данные
        for row_num, donation in enumerate(donations, 2):
            # ФИО донора
            sheet.cell(row=row_num, column=1, value=donation["donor_name"])

            # Дата сдачи крови
            date_str = donation["donation_date"].strftime("%d.%m.%Y") if donation["donation_date"] else ""
            sheet.cell(row=row_num, column=2, value=date_str)

            # Организатор (центр крови)
            sheet.cell(row=row_num, column=3, value=donation["organizer_name"])

        # Автоматическая ширина колонок
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True
            # Устанавливаем минимальную ширину
            sheet.column_dimensions[column_letter].width = max(20, sheet.column_dimensions[column_letter].width or 0)

        # Сохраняем в BytesIO
        file_content = BytesIO()
        workbook.save(file_content)
        file_content.seek(0)

        return file_content
